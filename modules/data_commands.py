from pandas import DataFrame
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pytz import timezone

try:
    from modules.sql_commands import sql_commands
except ModuleNotFoundError:
    from sql_commands import sql_commands

class data():
    # Создание таблиц
    def born_of_tables():
        sql_commands.change_in_table('bot', 'CREATE TABLE IF NOT EXISTS users (id int primary key, status varchar(50))')
        sql_commands.change_in_table('bot', 'CREATE TABLE IF NOT EXISTS bot_msg (msg_id int primary key, id int)')
        sql_commands.change_in_table('bot', 'CREATE TABLE IF NOT EXISTS pre_reg (id int primary key, name varchar(50), address varchar(100), phone varchar(12))')
        sql_commands.change_in_table('yabloko', 'CREATE TABLE IF NOT EXISTS voters (id int primary key, name varchar(50), address varchar(100), phone varchar(12))')

    # Удаление пред регистрационых данных
    def delete_pre_reg(id):
        sql_commands.change_in_table('bot', 'DELETE FROM pre_reg WHERE id == \'%d\'' % (id))

    # Удаление данных пользователя из бд
    def delete_voter_data(id):
        sql_commands.change_in_table('yabloko', 'DELETE FROM voters WHERE id == \'%d\'' % (id))

    # Информация об участке и кандидате округа
    def uchastok_info(street, house):
        izber_info = { 'uchastok' : '', 'candidate' : '' }

    # Преобразование текста в адрес
    def text_to_address(text):
        address = {
            'street' : '',
            'house' : '',
            'korp' : ''
        }

        text = text.split(',', 1)

        try:
            text = text[0] + ',' + text[1].replace(' ', '')
        except IndexError:
            return None

        text = text.lower()

        if data.check_text_address(text):
            text = text.split(',')
            address['street'] = text[0].title()
            address['house'] = text[1].replace('д.', '')
            if len(text) == 3:
                address['korp'] = text[2].replace('корп.', '')
        else:
            address = None

        return address

    # Проверка текста на прописание адреса
    def check_text_address(text):

        result = False

        if text.count(',') in [1,2]:
            text = text.split(',')
            if text[1].startswith('д.') and len(text[1]) > 2:
                if len(text) == 3:
                    if text[2].startswith('корп.') and len(text[2]) > 5:
                        result = True
                elif len(text) == 2:
                    result = True

        return result
    
    # Получение текста адреса из словаря
    def address_to_text(address):

        if address['korp'] == '':
            return '{}, д. {}'.format(address['street'] , address['house'])
        else:
            return '{}, д. {}, корп. {}'.format(address['street'] , address['house'], address['korp'])
    
    # Получение ексель таблицы с даными всех пользователей
    def get_excel():

        sql_data = sql_commands.grab_voters_info()

        buffer = BytesIO()

        date = timezone('Asia/Tomsk').localize(datetime.now()).astimezone(timezone('Europe/Moscow'))
        
        add_zero = lambda num: str(num).zfill(2)

        buffer.name = 'данные_пользователей {}-{}-2023 {}-{}.xlsx'.format(add_zero(date.day), add_zero(date.month), add_zero(date.hour), add_zero(date.minute))
        DataFrame(sql_data).to_excel(buffer, index=False)
        buffer.seek(0)

        workbook = load_workbook(buffer)

        sheet = workbook.active

        sheet['A1'] = 'Имя Фамилия'
        sheet['B1'] = 'Адрес'
        sheet['C1'] = 'Телефон'


        # Определение ширины столбцов на основе самой длинной записи в них
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    continue
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(buffer)

        buffer.seek(0)

        return buffer

if __name__ == '__main__':
    print(data.address_to_text(data.text_to_address('jrn, д.       123,Корп.2')))